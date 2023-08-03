import json
import os
from django.shortcuts import render, redirect
from django.http import HttpResponse
import openai
import eonchatapp
from eonchatapp import settings
from eonchatapp.settings import OPENAI_API_KEY
import pandas as pd

from django.templatetags.static import static

# excel file data importing and processing
excel_file_path = os.path.join(eonchatapp.settings.STATIC_ROOT, 'CBSE Syllabus-Class 9&10.xlsx')
# Read the excel file into a DataFrame
df = pd.read_excel(excel_file_path)

# Extract relevant information from the Excel data
df = df.stack().reset_index(drop=True).to_frame()

# Handle missing values
relevant_info = df.dropna()  # Drop rows with missing values

# Convert the extracted information to a string
relevant_info_str = ','.join(relevant_info.values.flatten())

''' ==================     try these other ways     ==================
Eon: Yes, there are multiple ways to convert Excel data into a processed format like a list or dataframe. Here are a few options:

1. Using pandas library: You can use the pandas library in Python to read the Excel file and convert it into a dataframe. Here's an example code snippet:

```python
import pandas as pd

# Read the Excel file into a dataframe
df = pd.read_excel('your_file.xlsx')

# Convert the dataframe into a list
data_list = df.values.tolist()
```

2. Using openpyxl library: Another option is to use the openpyxl library, which provides tools for working with Excel files. Here's an example code snippet:

```python
from openpyxl import load_workbook

# Load the Excel file
workbook = load_workbook(filename='your_file.xlsx')

# Select the specific sheet
sheet = workbook['Sheet1']

# Convert the data into a list
data_list = []
for row in sheet.iter_rows(values_only=True):
    data_list.append(list(row))
```

These are just a couple of examples, and there are other libraries and methods available depending on your specific requirements.

'''

syllabus_keywords = [
    "hi", "hello", "Phoenix Greens School", "CBSE syllabus", "Class", "Grade", "Maths", "Mathematics", "Science",
    "Physics", "Chemistry", "Biology", "History", "Economics", "Geography", "Politics", "English", "Social",
    "Telugu", "Hindi", "EVS", "Computer Science","Chapter", "Chapters", "Unit", "Units","Subject", "Subjects",
    "Topic", "Topics", "describe", "elaborate", "more", "above", "from", "OK", "Thanks", "cool",
    "great", "nice", "no", "yes"
    # Add more relevant keywords related to subjects, chapters, units, etc.
]

# Combine csv_syllabus_keywords and syllabus_keywords into a single list
#syllabus_keywords = csv_syllabus_keywords.split(',') + basic_keywords


# Variable to store the conversation history
conversation_history = []

def is_syllabus_related_question(question):
    question_lower = question.lower()
    for keyword in syllabus_keywords:
        if keyword.lower() in question_lower:
            return True
    return False


def get_custom_chatgpt_response(question):
    global conversation_history
    if is_syllabus_related_question(question):

        # API implementation
        # create openai instance
        openai.Model.list()
        # Set the API key -- Get OpenAI api key from environment variables path(eonchatapp-m3) done in settings.py file.
        openai.api_key = OPENAI_API_KEY
        #print(f"OPENAI_API_KEY: {OPENAI_API_KEY}")

        # Get the previous question and response from the conversation history
        # Get the previous question and response from the conversation history
        prev_question = conversation_history[-2] if len(conversation_history) >= 2 else ""
        prev_response = conversation_history[-1] if len(conversation_history) >= 1 else ""
        # print(prev_question+" : "+prev_response+"\n"+"***")

        # Construct the prompt using previous question and response
        prompt = f"{prev_question}\n{prev_response}"
        #print(prompt)

        # Append the current question to the conversation history
        conversation_history.append(question)

        # send relevant info as prompt to openai completion
        prompt_data = 'please provide information only related to '+relevant_info_str

        # Make a Completion
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": prompt_data},
                {"role": "assistant", "content": prompt},
                {"role": "user", "content": question}
            ],
            temperature=0.5,
            max_tokens=500,
            top_p=0,
            frequency_penalty=0,
            presence_penalty=0
        )
        # Check if response exists and has 'choices' list
        if response and 'choices' in response and response['choices']:
            response_content = response['choices'][0]['message']['content']
            response = response_content.strip() if response_content else 'Sorry, I could not answer your question. Please try again later or ask a different question.'
        else:
            response = 'Sorry, Something went wrong cannot process your request at the moment. Please try again later.'

        # Append the current question and AI response to the conversation history
        conversation_history.append(response)
        conversation_length = len(conversation_history)
        # print(conversation_history)
        return response, conversation_length

    elif not is_syllabus_related_question(question):
        # Respond to unrelated questions
        response = "Sorry, I can only answer questions related to the CBSE syllabus data for Phoenix Greens School. Please try asking a relevant question."
        conversation_history.append(question)
        conversation_history.append(response)
        conversation_length = len(conversation_history)
        return response, conversation_length

    else:
        quest=question
        res="Please ask a question to generate a response"
        conversation_history.append(quest)
        conversation_history.append(res)
        conversation_length = len(conversation_history)
        return res, conversation_length

def get_chatgpt_response(question):
    global conversation_history
    # API implementation
    # create openai instance
    openai.Model.list()
    # Set the API key -- Get OpenAI api key from environment variables path(eonchatapp-m3) done in settings.py file.
    openai.api_key = OPENAI_API_KEY
    # print(f"OPENAI_API_KEY: {OPENAI_API_KEY}")

    # Get the previous question and response from the conversation history
    # Get the previous question and response from the conversation history
    prev_question = conversation_history[-2] if len(conversation_history) >= 2 else ""
    prev_response = conversation_history[-1] if len(conversation_history) >= 1 else ""
    # print(prev_question+" : "+prev_response+"\n"+"***")

    # Construct the prompt using previous question and response
    prompt = f"{prev_question}\n{prev_response}"
    # print(prompt)
    prompt_data = "Answer as consisely as possible"
    # Append the current question to the conversation history
    conversation_history.append(question)
    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "system", "content": prompt_data},
            {"role": "assistant", "content": prompt},
            {"role": "user", "content": question}
        ],
        temperature=0.5,
        max_tokens=500,
        top_p=0,
        frequency_penalty=0,
        presence_penalty=0
    )
    # Check if response exists and has 'choices' list
    if response and 'choices' in response and response['choices']:
        response_content = response['choices'][0]['message']['content']
        response = response_content.strip() if response_content else 'Sorry, I could not answer your question. Please try again later or ask a different question.'
    else:
        response = 'Sorry, Something went wrong cannot process your request at the moment. Please try again later.'

    # Append the current question and AI response to the conversation history
    conversation_history.append(response)
    conversation_length = len(conversation_history)
    # print(conversation_history)

    return response, conversation_length

def response_view(request):
    # Your code to generate the context for the response_view.html template
    # For example:
    if request.method == "POST":
        question = request.POST.get('question')
        response, conversation_length = get_chatgpt_response(question)
        return render(request, "response_view.html", {"question": question, "response": response, "conversation_history": conversation_history})

    return render(request, "response_view.html", {})

def home(request):
    # Check for form submission
    if request.method == "POST":
        question = request.POST.get('question')

        # Get the value of the toggle switch
        print("Form data:", request.POST)

        toggle_switch = request.POST.get('toggle_switch_checked')
        if toggle_switch == 'on':
            response, conversation_length = get_custom_chatgpt_response(question)
        else :
            response, conversation_length = get_chatgpt_response(question)


        return render(request, "home.html", {"question": question, "response": response, "conversation_history": conversation_history})

    return render(request, "home.html", {})

